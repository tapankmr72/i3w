import streamlit as st 
import time
from datetime import datetime
import json
import urllib.request
import requests
import openpyxl
path=""
wb_obj1 = openpyxl.load_workbook(path + "mapped.xlsx")
sheet_obja = wb_obj1.active
wb_obj = openpyxl.load_workbook(path+"latlong.xlsx")
sheet_obj = wb_obj.active

polltime=2
pollgap=2
u2=""
replyto="5560841599"
looper=0
message = ""
updatetext1=0
#token="6277980563:AAEIvKjZkwvNNA_jhq_Ur_SQTJfJkgETTb4"
headers = {"accept": "application/json","content-type": "application/json"}
photourl="https://api.telegram.org/bot"+token+"/sendPhoto"
docurl="https://api.telegram.org/bot"+token+"/sendDocument"
texturl="https://api.telegram.org/bot"+token+"/sendMessage"
meurl ="https://api.telegram.org/bot"+token+"/getMe"
pollurl="https://api.telegram.org/bot"+token+"/getUpdates"
filedownload="https://api.telegram.org/file/bot"+token+"/"
fileurl="https://api.telegram.org/bot"+token+"/getfile"
healthduration=30
#in minutes
healthmessage="This is health message of Basic Telebot. It is running fine and you are receiving this message every"+ str(healthduration) +"  minutes "
healthtime = int(time.time())

while looper==0:
    message=""
    alert=""
    cn = 0
    filename = ""
    mime=""
    messagelink = ""
    updatefile = open(path + "updateid.txt", 'r+')
    updatetext = updatefile.read()
    updatefile.close()
    #Starting Long Polling
    lastupdate=int(updatetext)
    payloadpoll = {"offset":lastupdate,"limit": 50,"timeout": pollgap}
    response = requests.post(pollurl, json=payloadpoll, headers=headers)
    f = urllib.request.urlopen(pollurl)
    data=json.load(f)
    data=str(data)
    lendata=len(data)
    print(data)
    #print(lendata)
    if lendata<50:
        print("No New message "+str(datetime.now().strftime("%H:%M:%S")))
    if lendata>50:
      print("Message received at: "+str(datetime.now().strftime("%H:%M:%S")))
      tempstr = data[0:lendata]
      latid=""
      longid=""
      while cn==0:
       downloadlink=""
       usernametext=""
       usernamepos1=tempstr.find("username")
       if usernamepos1!=-1:
         usernamepos2=tempstr.find("language_code")
         usernametext=tempstr[usernamepos1+12:usernamepos2-4]
       locationpos = tempstr.find("location': {'latitude':")
       if locationpos != -1:
           latid = tempstr.find("latitude")
           longid = tempstr.find("longitude")
           livepos = tempstr.find("live_period")
           accuracypos=tempstr.find("horizontal_accuracy")
           latid = tempstr[latid + 11:longid-2]
           latid = latid.strip()
           latid=latid.strip(",")
           if livepos != -1:
            longid = tempstr[longid + 11:livepos-2]
            longid = longid.strip()
            longid = longid.rstrip(",")

           else:

            if accuracypos!=-1:
                longid = tempstr[longid + 11:accuracypos-2]
            else:
               findlastlongid=tempstr.find("}}}]}")
               longid = tempstr[longid + 11:findlastlongid]

            #longid = longid.rstrip("}")
            longid = longid.strip()
            longid = longid.rstrip(",")
       print("latlong",latid,longid)
       filepos1 = tempstr.rfind("file_id':")
       if filepos1 != -1:
           filenamepos1 = tempstr.find("file_name':")
           filenamepos2 = tempstr.find("mime_type':")
           filename = tempstr[filenamepos1 + 13:filenamepos2 - 4]
           print("Uploaded File name :" + filename)
           filepos2 = tempstr.rfind("'file_unique_id':")
           fileid = tempstr[filepos1 + 11:filepos2 - 3]
           #print(fileid)
           payload = {"file_id": fileid}
           response = requests.post(fileurl, json=payload, headers=headers)
           ok=response.text
           #print(response.text)
           if  ok.find("true")!=-1:
               linkpos1 = ok.find("file_path")
               linkpos2 = ok.find("}}")
               link=ok[linkpos1 + 12:linkpos2 - 1]
               print(link)
               finddot= link.rfind(".")
               if finddot!=-1:
                   mime=link[finddot:len(link)]
                   #print("mime:"+mime)
                   if mime != ".jpg" and mime != ".png" and mime != ".jpeg" and mime != ".bmp" and mime != ".webp" and mime =="":
                       messagelink="This is not a valid Image file. Please check and send Image file only"
               downloadlink=filedownload+link
               #print(downloadlink)
       messagepostemp=0
       updatepos1 = tempstr.find("update_id':")
       #print("updatepos1",updatepos1)
       if updatepos1==-1:
        break
       namepos = tempstr.find("is_bot': False, 'first_name':")
       namepos1 = tempstr.find("last_name")
       namepos2 = tempstr.find("username")
       namepos3 = tempstr.find("language_code")
       #print("namepos", namepos)
       #print("namepos1", namepos1)
       numberpos = tempstr.find("from")
       #print("numberpos",numberpos)
       datepos = tempstr.find("date':")
       #print("datepos",datepos)
       endpos1 = tempstr.find("}}, {")
       #print("endpos1",endpos1)
       endpos2 = tempstr.find("}}]}")
       messagepos = tempstr.find("text")
       if messagepos-datepos>20:
         messagepostemp =messagepos
         messagepos=datepos+20
       #print("messagepos",messagepos)
       updatetext1 = tempstr[updatepos1 + 12:updatepos1 + 21]
       updatetext1=updatetext1.rstrip(",")
       print(updatetext1)
       numbertext = tempstr[numberpos + 14:numberpos + 24]
       numbertext = numbertext.strip()
       numbertext = numbertext.rstrip(",")
       #print(numbertext)
       #print(namepos1)


       if namepos1 == -1 :
           nametext = tempstr[namepos + 31:namepos3 - 4]

       if namepos1 != -1:
           nametext = tempstr[namepos + 31:namepos1 - 4]

       if namepos1 == -1 and namepos2!=-1 :
           nametext = tempstr[namepos + 31:namepos2 - 4]

       if namepos1 == -1 and namepos2 == -1 and namepos3!=-1:
           nametext = tempstr[namepos + 31:namepos3 - 4]

       #print("Name: " + nametext)
       datetext = tempstr[datepos + 7:datepos + 17]
       dateconfirm = datetext[2:10]
       datetext=(datetime.fromtimestamp(int(datetext)).strftime('%Y-%m-%d %H:%M:%S'))
       #print(datetext)
       if endpos1!=-1:
        messagetext=tempstr[messagepos+8:endpos1-1]
        if messagepostemp-datepos>20:
          messagetext="No Text"
       if endpos1==-1:
        messagetext=tempstr[messagepos+8:endpos2-1]
        if messagepostemp-datepos>20:
          messagetext="No Text"
       #print(messagetext)
       messagetext = messagetext.replace("\\xa0", " ")

       #print(messagetext)
       #print("----------------")

       #print("username: "+usernametext)

       apos=nametext.find(",")
       if apos!=-1:
           u1=nametext[0:apos-1]
       else:
        u1 = nametext

       if u1 != u2:
           alert = u1 + " has just logged in\n\n"
           alert1 = "<a href='tg://user?id=" + numbertext + "'>Click to chat here</a>"
           if usernametext!="":
               alert1=alert1+"\n\nor "+"@"+usernametext
           payloadtext = {"text": alert +alert1, "parse_mode": "html", "disable_web_page_preview": False,
                          "disable_notification": False, "reply_to_message_id": None, "chat_id": replyto}
           response = requests.post(texturl, json=payloadtext, headers=headers)
           u2 = u1
       callbackpos=messagetext.find("'data':")
       callbacktext=messagetext[callbackpos+9:len(messagetext)]
       #print("callbacktext:",callbacktext)
       numbertext=numbertext.rstrip(",")
       spcfind = messagetext.find("entities")

       if spcfind != -1:
           messagetext = messagetext[0:spcfind - 4]
       print("Logged user:" + alert)
       print(alert1)
       print("-----------------------------------------------------------")
       print("USER ID: ",numbertext)
       print("Date: ",datetext)
       print("Name: " + nametext)
       print("username@: " + usernametext)
       print("User First Name: " + u1)

       print("callbacktext: ", callbacktext)
       print("Downloadlink: "+downloadlink)
       print("Message: "+messagetext)
       print("Latitude: " + latid)
       print("Longitude: " + longid)
       print("mime:" + mime)
       print("Uploaded File name : " + filename)
       print("-----------------------------------------------------------")

       if latid!="":
           with open("latlongtoplace.py") as f:
               exec(f.read())
               if address=="":
                   address="Shared location is not within India.."


               payloadtext = {"text": address+"\n\n"+lat+"\n"+long, "parse_mode": "html",
                              "disable_web_page_preview": True,
                              "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
               response = requests.post(texturl, json=payloadtext, headers=headers)
               tempstr = tempstr[datepos + 65:lendata]
               # sr=1
               break

       elif messagetext[0:3].upper()=="MAP":
           findto=messagetext.upper().find("TO")
           mapfrom=messagetext[3:findto].strip()
           mapto = messagetext[findto+2:len(messagetext)].strip()
           print(mapfrom)
           print(mapto)
           wb_obj1 = openpyxl.load_workbook(path + "mapped.xlsx")
           sheet_obja = wb_obj1.active
           n = 0
           srow = 2
           while n == 0:
               cell_obj1a = sheet_obja.cell(row=srow, column=1)
               cell_obj2a = sheet_obja.cell(row=srow, column=2)
               if cell_obj1a.value ==None or cell_obj1a.value =="":
                   cell_obj1a.value = mapfrom
                   cell_obj2a.value = mapto
                   wb_obj1.save(path + "mapped.xlsx")
                   message = "Mapping done successfully"
                   payloadtext = {"text": message, "parse_mode": "html",
                                  "disable_web_page_preview": False,
                                  "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
                   response = requests.post(texturl, json=payloadtext, headers=headers)
                   break
               else:
                srow=srow+1


       else:
           text = messagetext.upper()
           strLen = len(text)
           range_obj = range(strLen)
           character = "."
           count = 0
           count1 = ""
           for index in range_obj:
               if text[index] == character:
                   count = count + 1
                   count1 = count1 + "<" + str(index)
           print(count1)
           print(count)
           if count != 2:
               print("not proper string")
               message = "?"
               payloadtext = {"text": message, "parse_mode": "html",
                              "disable_web_page_preview": False,
                              "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
               response = requests.post(texturl, json=payloadtext, headers=headers)
               break

           with open("decodelatlong.py") as f:
               exec(f.read())

               payloadtext = {"text":message, "parse_mode": "html",
                              "disable_web_page_preview": False,
                              "disable_notification": False, "reply_to_message_id": None, "chat_id": numbertext}
               response = requests.post(texturl, json=payloadtext, headers=headers)
               tempstr = tempstr[datepos + 65:lendata]
               # sr=1
               break



       break
       tempstr = tempstr[endpos1+4 :lenc]


    if int(updatetext1)>=int(updatetext):
      updatefile = open(path + "updateid.txt", 'w')
      updatetext = int(updatetext1)+1
      updatefile.write(str(updatetext))
      updatefile.close()

    healthtime1 = int(time.time())
    # print(healthtime1)
    if healthtime1 - healthtime > 1800:
        payloadtext = {"text": healthmessage, "parse_mode": "html", "disable_web_page_preview": False,
                       "disable_notification": False, "reply_to_message_id": None, "chat_id": replyto}
        response = requests.post(texturl, json=payloadtext, headers=headers)
        healthtime = healthtime1

    print("Offset ID in Text File updated succesfully")
